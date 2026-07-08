import logging
import multiprocessing
import shutil
import tempfile
from concurrent.futures import ProcessPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QThread, Signal

from src.core.evolution_analyzer import EvolutionAnalyzer
from src.core.io_sync import PipelineIO
from src.core.physics import CrackPhysicsEngine

logger = logging.getLogger(__name__)

_worker_engine = None


@dataclass(frozen=True)
class FrameTaskPayload:
    config: dict
    u_path: str
    exx_path: str
    mask_path: str
    v_path: Optional[str]
    quality_path: Optional[str]
    ratio: float
    dic_point_spacing_mm: float
    subset_spacing_px: float
    metadata_source: str
    frame_id: int
    time_s: float
    time_source: str


def _median_band(arr: np.ndarray, valid: np.ndarray, x_center: int, half_width: int) -> float:
    x0 = max(0, x_center - half_width)
    x1 = min(arr.shape[1], x_center + half_width + 1)
    roi = arr[:, x0:x1]
    roi_valid = valid[:, x0:x1] & np.isfinite(roi)
    if np.count_nonzero(roi_valid) == 0:
        return np.nan
    return float(np.nanmedian(roi[roi_valid]))


def _virtual_extensometer_strain(
    u: np.ndarray,
    valid_mask: np.ndarray,
    displacement_scale_mm: float,
    dic_point_spacing_mm: float,
    config: dict,
) -> tuple[float, float, int, int, str]:
    experiment = config.get("experiment", {})
    ve = experiment.get("virtual_extensometer", {})
    active_cols = np.where(np.any(valid_mask & np.isfinite(u), axis=0))[0]
    if active_cols.size < 5:
        return 0.0, float(experiment.get("gauge_length_mm", 80.0)), -1, -1, "no_valid_columns"

    left_fraction = float(ve.get("left_fraction", 0.10))
    right_fraction = float(ve.get("right_fraction", 0.90))
    band_width_points = max(1, int(ve.get("band_width_points", 3)))
    x_min, x_max = int(active_cols[0]), int(active_cols[-1])
    span = max(1, x_max - x_min)
    left_x = int(round(x_min + span * left_fraction))
    right_x = int(round(x_min + span * right_fraction))
    if right_x <= left_x:
        return 0.0, float(experiment.get("gauge_length_mm", 80.0)), left_x, right_x, "invalid_extensometer"

    left_u = _median_band(u, valid_mask, left_x, band_width_points)
    right_u = _median_band(u, valid_mask, right_x, band_width_points)
    if np.isnan(left_u) or np.isnan(right_u):
        return 0.0, float(experiment.get("gauge_length_mm", 80.0)), left_x, right_x, "missing_band_displacement"

    configured_l0 = ve.get("gauge_length_mm")
    if configured_l0 is not None and float(configured_l0) > 0:
        gauge_len = float(configured_l0)
        source = "dic_virtual_extensometer_configured_L0"
    else:
        gauge_len = abs(right_x - left_x) * float(dic_point_spacing_mm)
        source = "dic_virtual_extensometer_grid_L0"

    if gauge_len <= 0:
        return 0.0, float(experiment.get("gauge_length_mm", 80.0)), left_x, right_x, "invalid_gauge_length"
    du_mm = abs(right_u - left_u) * float(displacement_scale_mm)
    return max(0.0, float(du_mm / gauge_len)), gauge_len, left_x, right_x, source


def analyze_single_frame_task(payload: FrameTaskPayload) -> Optional[Dict[str, Any]]:
    global _worker_engine
    try:
        if _worker_engine is None:
            _worker_engine = CrackPhysicsEngine(payload.config)

        u = np.load(payload.u_path)
        exx = np.load(payload.exx_path)
        mask = np.load(payload.mask_path).astype(bool)
        v = np.load(payload.v_path) if payload.v_path else None
        quality = np.load(payload.quality_path) if payload.quality_path else None

        quality_mask, quality_fraction, quality_reason = _worker_engine.build_quality_mask(mask, u, v, quality)
        strain, virtual_l0, left_x, right_x, strain_source = _virtual_extensometer_strain(
            u, quality_mask, payload.ratio, payload.dic_point_spacing_mm, payload.config
        )

        if quality_fraction < _worker_engine.min_valid_fraction:
            res = _worker_engine._empty("quality_rejected")
            threshold = 0.0
        else:
            skeleton, threshold = _worker_engine.extract_skeleton(exx, quality_mask)
            res = _worker_engine.compute_cod(
                u,
                skeleton,
                payload.ratio,
                payload.dic_point_spacing_mm,
                v_map=v,
                sample_mask=quality_mask,
            )

        res.update(
            {
                "Frame": int(payload.frame_id),
                "Time_s": float(payload.time_s),
                "dic_time_source": payload.time_source,
                "global_strain": max(0.0, float(strain)),
                "virtual_gauge_length_mm": float(virtual_l0),
                "virtual_left_col": int(left_x),
                "virtual_right_col": int(right_x),
                "strain_source": strain_source,
                "quality_valid_fraction": float(quality_fraction),
                "quality_filter": quality_reason,
                "quality_map_present": bool(quality is not None),
                "strain_threshold_used": float(threshold),
                "pixel_size_mm": float(payload.ratio),
                "subset_spacing_px": float(payload.subset_spacing_px),
                "dic_point_spacing_mm": float(payload.dic_point_spacing_mm),
                "metadata_source": payload.metadata_source,
                "v_map_present": bool(v is not None),
            }
        )

        crack_count = int(res.get("crack_count", 0))
        res["crack_spacing_mm"] = float(virtual_l0 / crack_count) if crack_count > 0 and virtual_l0 > 0 else 0.0

        return res
    except Exception as e:
        logger.error("Frame task crashed (Frame %s): %s", payload.frame_id, e, exc_info=True)
        return None


class AnalysisPipelineWorker(QThread):
    progress_updated = Signal(int, int)
    log_emitted = Signal(str)
    error_occurred = Signal(str)
    finished = Signal()
    specimen_processed = Signal(str, str)

    def __init__(self, paired_data: dict, out_dir: Path, config: dict) -> None:
        super().__init__()
        self.paired_data = paired_data
        self.out_dir = out_dir
        self.config = config
        self._is_running = True

    def run(self) -> None:
        try:
            experiment = self.config.get("experiment", {})
            fallback_ratio = float(experiment.get("mm_per_pixel", 0.045))
            interval = float(experiment.get("sampling_interval_s", 5.0))
            if interval <= 0:
                raise ValueError("experiment.sampling_interval_s must be greater than zero.")
            total = len(self.paired_data)

            for i, (mat_f, mts_f) in enumerate(self.paired_data.items()):
                if not self._is_running:
                    break
                try:
                    self._process_specimen(Path(mat_f), Path(mts_f) if mts_f else None, fallback_ratio, interval)
                except Exception as e:
                    self.log_emitted.emit(f"Failed specimen {Path(mat_f).name}; skipped: {e}")
                    logger.error("Specimen process failed: %s", mat_f, exc_info=True)
                self.progress_updated.emit(i + 1, total)

        except Exception as e:
            self.error_occurred.emit(f"Pipeline scheduler crashed: {e}")
        finally:
            self.finished.emit()

    @staticmethod
    def _resolve_frame_time(frame: Any, fallback_interval_s: float) -> tuple[float, str]:
        raw_time = getattr(frame, "time_s", np.nan)
        if raw_time is not None and np.isfinite(raw_time):
            return float(raw_time), "mat_metadata_time"
        return float(frame.frame_id * fallback_interval_s), "frame_index_interval_fallback"

    def _process_specimen(self, mat_path: Path, mts_path: Optional[Path], ratio: float, interval: float) -> None:
        temp_dir = Path(tempfile.mkdtemp(prefix="cv_"))
        tasks = []
        first_meta = None

        try:
            for frame in PipelineIO.stream_dic_frames(mat_path, ratio, self.config):
                if first_meta is None:
                    first_meta = frame
                    self.log_emitted.emit(
                        f"Parsing {mat_path.name} | pixel={frame.ratio:.6f} mm/px | "
                        f"DIC step={frame.subset_spacing_px:.3f} px | "
                        f"grid={frame.dic_point_spacing_mm:.6f} mm/point | source={frame.metadata_source}"
                    )

                frame_time_s, time_source = self._resolve_frame_time(frame, interval)

                u_p = temp_dir / f"u_{frame.frame_id}.npy"
                exx_p = temp_dir / f"exx_{frame.frame_id}.npy"
                mask_p = temp_dir / f"mask_{frame.frame_id}.npy"
                v_p = temp_dir / f"v_{frame.frame_id}.npy" if frame.v_map is not None else None
                q_p = temp_dir / f"q_{frame.frame_id}.npy" if frame.quality_map is not None else None
                np.save(u_p, frame.u_map)
                np.save(exx_p, frame.exx_map)
                np.save(mask_p, frame.mask)
                if v_p is not None:
                    np.save(v_p, frame.v_map)
                if q_p is not None:
                    np.save(q_p, frame.quality_map)

                tasks.append(
                    FrameTaskPayload(
                        self.config,
                        str(u_p),
                        str(exx_p),
                        str(mask_p),
                        str(v_p) if v_p is not None else None,
                        str(q_p) if q_p is not None else None,
                        frame.ratio,
                        float(frame.dic_point_spacing_mm),
                        frame.subset_spacing_px,
                        frame.metadata_source,
                        frame.frame_id,
                        frame_time_s,
                        time_source,
                    )
                )

            if not tasks:
                self.log_emitted.emit(f"{mat_path.name} contains no DIC frames.")
                return

            results = []
            max_workers = min(10, max(1, multiprocessing.cpu_count() - 2))

            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                for res in executor.map(analyze_single_frame_task, tasks):
                    if res:
                        results.append(res)

            if not results:
                self.log_emitted.emit(f"{mat_path.name} extraction failed: no valid frames.")
                return

            results.sort(key=lambda x: (x["Time_s"], x["Frame"]))
            if bool(self.config.get("physics", {}).get("enforce_monotonic_strain", True)):
                cur_max_strain = 0.0
                for res in results:
                    cur_max_strain = max(cur_max_strain, res.get("global_strain", 0.0))
                    res["global_strain"] = cur_max_strain

            df = pd.DataFrame(results)

            if df["Time_s"].duplicated().any():
                dup = int(df["Time_s"].duplicated().sum())
                self.log_emitted.emit(f"⚠️ {mat_path.name}: detected {dup} duplicated DIC timestamps; MTS sync may be rejected.")

            if mts_path and mts_path.exists():
                try:
                    df = EvolutionAnalyzer(self.config, mts_path).synchronize(df)
                    self.log_emitted.emit(f"{mat_path.name} MTS sync passed strict overlap checks.")
                except Exception as e:
                    df["sync_status"] = f"failed: {e}"
                    self.log_emitted.emit(f"MTS sync failed; kept DIC virtual strain: {e}")
            else:
                df["sync_status"] = "no_mts"

            self._export_results(mat_path, df, results)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _export_results(self, mat_path: Path, df: pd.DataFrame, results: list[Dict[str, Any]]) -> None:
        specimen = mat_path.stem
        frame_df = self._prepare_frame_table(specimen, df)
        key_rows = self._select_key_rows(frame_df)
        sat_row = key_rows["Saturated"]
        ult_row = key_rows["Ultimate"]

        target_df = self._build_target_state_table(specimen, frame_df)
        key_crack_df = self._build_key_crack_table(specimen, results, key_rows)
        target_crack_df = self._build_target_crack_table(specimen, results, target_df)
        crack_tidy_df = pd.concat([key_crack_df, target_crack_df], ignore_index=True)
        distribution_df = self._build_distribution_table(crack_tidy_df)
        summary_df = self._build_specimen_summary(specimen, frame_df, sat_row, ult_row, crack_tidy_df)
        qa_frame_df = self._build_qa_frame_status(frame_df)
        qa_meta_df = self._qa_frame(frame_df)
        validation_df = self._validation_frame(mat_path, frame_df)
        guide_df = self._sheet_guide_frame()

        origin_f = self.out_dir / f"{specimen}_Origin_Plot_Data.xlsx"
        with pd.ExcelWriter(origin_f, engine="openpyxl") as writer:
            self._write_sheet(writer, guide_df, "00_READ_ME")
            self._write_sheet(writer, self._origin_curve_table(frame_df), "01_Frame_Curves")
            self._write_sheet(writer, target_df, "02_Target_States")
            self._write_sheet(writer, distribution_df, "03_Distribution_Tidy")
            self._write_sheet(writer, crack_tidy_df, "04_Crack_Tidy")

        stat_f = self.out_dir / f"{specimen}_Statistics_Report.xlsx"
        with pd.ExcelWriter(stat_f, engine="openpyxl") as writer:
            self._write_sheet(writer, summary_df, "00_Specimen_Summary")
            self._write_sheet(writer, frame_df, "01_Frame_All")
            self._write_sheet(writer, target_df, "02_Target_Summary")
            self._write_sheet(writer, key_crack_df, "03_Key_Crack_Details")
            self._write_sheet(writer, distribution_df, "04_Distribution_Tidy")
            self._write_sheet(writer, qa_frame_df, "05_QA_Frame_Status")
            self._write_sheet(writer, qa_meta_df, "06_QA_Metadata")
            self._write_sheet(writer, validation_df, "07_Validation")

        self._update_batch_summary(summary_df, target_df)

        for r in results:
            r.pop("per_crack_details", None)
            r.pop("raw_widths", None)

        self.log_emitted.emit(f"Origin data generated: {origin_f.name}")
        self.log_emitted.emit(f"Statistics report generated: {stat_f.name}")
        self.specimen_processed.emit(str(origin_f), str(stat_f))

    @staticmethod
    def _prepare_frame_table(specimen: str, df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        out["Specimen"] = specimen
        out["Strain_pct"] = pd.to_numeric(out.get("global_strain", 0.0), errors="coerce").fillna(0.0) * 100.0
        out["W_avg_um"] = pd.to_numeric(out.get("w_avg", 0.0), errors="coerce").fillna(0.0) * 1000.0
        out["W_max_um"] = pd.to_numeric(out.get("w_max", 0.0), errors="coerce").fillna(0.0) * 1000.0
        out["W_99_um"] = pd.to_numeric(out.get("w_99", 0.0), errors="coerce").fillna(0.0) * 1000.0
        out["crack_spacing_mm"] = pd.to_numeric(out.get("crack_spacing_mm", 0.0), errors="coerce").fillna(0.0)

        for col in [
            "Frame",
            "Time_s",
            "global_strain",
            "Strain_pct",
            "Stress_MPa",
            "Force_N",
            "Disp_mm",
            "MTS_Strain",
            "crack_count",
            "crack_spacing_mm",
            "W_avg_um",
            "W_99_um",
            "W_max_um",
            "cod_sample_count",
            "quality_valid_fraction",
            "strain_threshold_used",
            "pixel_size_mm",
            "subset_spacing_px",
            "dic_point_spacing_mm",
        ]:
            if col not in out.columns:
                out[col] = np.nan

        ult_s_raw = float(pd.to_numeric(out["global_strain"], errors="coerce").max(skipna=True) or 0.0)
        out["Normalized_Strain"] = out["global_strain"] / ult_s_raw if ult_s_raw > 0 else 0.0
        time_span = float(pd.to_numeric(out["Time_s"], errors="coerce").max(skipna=True) or 0.0)
        out["Normalized_Time"] = out["Time_s"] / time_span if time_span > 0 else 0.0

        drop_cols = ["raw_widths", "per_crack_details"]
        out = out.drop(columns=[c for c in drop_cols if c in out.columns], errors="ignore")

        front_cols = [
            "Specimen",
            "Frame",
            "Time_s",
            "Normalized_Time",
            "global_strain",
            "Strain_pct",
            "Normalized_Strain",
            "Stress_MPa",
            "Force_N",
            "Disp_mm",
            "MTS_Strain",
            "crack_count",
            "crack_spacing_mm",
            "W_avg_um",
            "W_99_um",
            "W_max_um",
            "cod_sample_count",
            "quality_valid_fraction",
            "strain_threshold_used",
            "cod_status",
            "sync_status",
            "dic_time_source",
            "strain_source",
            "metadata_source",
            "v_map_present",
            "quality_map_present",
            "quality_filter",
            "cod_vector_mode",
            "pixel_size_mm",
            "subset_spacing_px",
            "dic_point_spacing_mm",
            "virtual_gauge_length_mm",
            "virtual_left_col",
            "virtual_right_col",
        ]
        ordered_cols = [c for c in front_cols if c in out.columns] + [c for c in out.columns if c not in front_cols]
        return out[ordered_cols].sort_values(["Time_s", "Frame"], na_position="last").reset_index(drop=True)

    @staticmethod
    def _select_key_rows(frame_df: pd.DataFrame) -> dict[str, pd.Series]:
        if frame_df.empty:
            empty = pd.Series(dtype=object)
            return {"Saturated": empty, "Ultimate": empty, "First_Crack": empty, "Max_Width": empty}

        crack_count = pd.to_numeric(frame_df["crack_count"], errors="coerce").fillna(0.0)
        sat_idx = crack_count.idxmax()

        if "Stress_MPa" in frame_df.columns and not frame_df["Stress_MPa"].isna().all():
            ult_idx = pd.to_numeric(frame_df["Stress_MPa"], errors="coerce").idxmax()
        else:
            ult_idx = pd.to_numeric(frame_df["Strain_pct"], errors="coerce").idxmax()

        width_idx = pd.to_numeric(frame_df["W_max_um"], errors="coerce").fillna(0.0).idxmax()
        opened = frame_df[crack_count > 0]
        first_idx = opened.index[0] if not opened.empty else sat_idx

        return {
            "Saturated": frame_df.loc[sat_idx],
            "Ultimate": frame_df.loc[ult_idx],
            "First_Crack": frame_df.loc[first_idx],
            "Max_Width": frame_df.loc[width_idx],
        }

    def _build_target_state_table(self, specimen: str, frame_df: pd.DataFrame) -> pd.DataFrame:
        rows = []
        targets = self.config.get("export", {}).get("target_strains", [0.2, 2.0, 4.0, 6.0])
        max_strain = float(pd.to_numeric(frame_df["Strain_pct"], errors="coerce").max(skipna=True) or 0.0)

        for target in targets:
            target = float(target)
            if frame_df.empty:
                rows.append({"Specimen": specimen, "State": f"Target_{target:g}%", "Target_Strain_pct": target, "Status": "empty"})
                continue

            if target <= max_strain:
                idx = (pd.to_numeric(frame_df["Strain_pct"], errors="coerce") - target).abs().idxmin()
                row = frame_df.loc[idx]
                status = "reached"
            else:
                idx = pd.to_numeric(frame_df["Strain_pct"], errors="coerce").idxmax()
                row = frame_df.loc[idx]
                status = "not_reached"

            rows.append(self._state_row(specimen, f"Target_{target:g}%", row, target, status))

        return pd.DataFrame(rows)

    @staticmethod
    def _state_row(specimen: str, state: str, row: pd.Series, target: float | None, status: str) -> dict[str, Any]:
        return {
            "Specimen": specimen,
            "State": state,
            "Target_Strain_pct": target,
            "Status": status,
            "Frame": int(row.get("Frame", -1)) if pd.notna(row.get("Frame", np.nan)) else np.nan,
            "Time_s": float(row.get("Time_s", np.nan)) if pd.notna(row.get("Time_s", np.nan)) else np.nan,
            "Real_Strain_pct": float(row.get("Strain_pct", np.nan)) if pd.notna(row.get("Strain_pct", np.nan)) else np.nan,
            "Normalized_Strain": float(row.get("Normalized_Strain", np.nan)) if pd.notna(row.get("Normalized_Strain", np.nan)) else np.nan,
            "Stress_MPa": float(row.get("Stress_MPa", np.nan)) if pd.notna(row.get("Stress_MPa", np.nan)) else np.nan,
            "Crack_Count": int(row.get("crack_count", 0)) if pd.notna(row.get("crack_count", np.nan)) else 0,
            "Crack_Spacing_mm": float(row.get("crack_spacing_mm", np.nan)) if pd.notna(row.get("crack_spacing_mm", np.nan)) else np.nan,
            "W_avg_um": float(row.get("W_avg_um", np.nan)) if pd.notna(row.get("W_avg_um", np.nan)) else np.nan,
            "W_99_um": float(row.get("W_99_um", np.nan)) if pd.notna(row.get("W_99_um", np.nan)) else np.nan,
            "W_max_um": float(row.get("W_max_um", np.nan)) if pd.notna(row.get("W_max_um", np.nan)) else np.nan,
            "Quality_Valid_Fraction": float(row.get("quality_valid_fraction", np.nan)) if pd.notna(row.get("quality_valid_fraction", np.nan)) else np.nan,
            "COD_Status": str(row.get("cod_status", "")),
            "Sync_Status": str(row.get("sync_status", "")),
        }

    def _build_key_crack_table(
        self, specimen: str, results: list[Dict[str, Any]], key_rows: dict[str, pd.Series]
    ) -> pd.DataFrame:
        tables = []
        for state, row in key_rows.items():
            if row.empty:
                continue
            details = self._details_for_frame(results, row.get("Frame"))
            tables.append(self._format_crack_details(specimen, state, None, row, details))
        if tables:
            return pd.concat(tables, ignore_index=True)
        return self._empty_crack_table()

    def _build_target_crack_table(self, specimen: str, results: list[Dict[str, Any]], target_df: pd.DataFrame) -> pd.DataFrame:
        tables = []
        for _, state in target_df.iterrows():
            if state.get("Status") != "reached":
                continue
            frame = state.get("Frame")
            details = self._details_for_frame(results, frame)
            row = pd.Series(
                {
                    "Frame": frame,
                    "Time_s": state.get("Time_s"),
                    "Strain_pct": state.get("Real_Strain_pct"),
                    "Stress_MPa": state.get("Stress_MPa"),
                    "quality_valid_fraction": state.get("Quality_Valid_Fraction"),
                    "cod_status": state.get("COD_Status"),
                }
            )
            tables.append(self._format_crack_details(specimen, state.get("State", "Target"), state.get("Target_Strain_pct"), row, details))
        if tables:
            return pd.concat(tables, ignore_index=True)
        return self._empty_crack_table()

    @staticmethod
    def _empty_crack_table() -> pd.DataFrame:
        return pd.DataFrame(
            columns=[
                "Specimen",
                "State",
                "Target_Strain_pct",
                "Frame",
                "Time_s",
                "Real_Strain_pct",
                "Stress_MPa",
                "Crack_ID",
                "Length_mm",
                "W_avg_um",
                "W_max_um",
                "Sample_Count",
                "Quality_Valid_Fraction",
                "COD_Status",
            ]
        )

    @classmethod
    def _format_crack_details(
        cls, specimen: str, state: str, target: float | None, frame_row: pd.Series, details: pd.DataFrame
    ) -> pd.DataFrame:
        if details is None or details.empty:
            return cls._empty_crack_table()

        out = details.copy()
        out["Specimen"] = specimen
        out["State"] = state
        out["Target_Strain_pct"] = target
        out["Frame"] = int(frame_row.get("Frame", -1)) if pd.notna(frame_row.get("Frame", np.nan)) else np.nan
        out["Time_s"] = float(frame_row.get("Time_s", np.nan)) if pd.notna(frame_row.get("Time_s", np.nan)) else np.nan
        out["Real_Strain_pct"] = float(frame_row.get("Strain_pct", np.nan)) if pd.notna(frame_row.get("Strain_pct", np.nan)) else np.nan
        out["Stress_MPa"] = float(frame_row.get("Stress_MPa", np.nan)) if pd.notna(frame_row.get("Stress_MPa", np.nan)) else np.nan
        out["W_avg_um"] = pd.to_numeric(out.get("W_avg_mm", np.nan), errors="coerce") * 1000.0
        out["W_max_um"] = pd.to_numeric(out.get("W_max_mm", np.nan), errors="coerce") * 1000.0
        out["Length_mm"] = pd.to_numeric(out.get("Length_mm", out.get("L_mm", np.nan)), errors="coerce")
        out["Sample_Count"] = pd.to_numeric(out.get("count", np.nan), errors="coerce")
        out["Quality_Valid_Fraction"] = frame_row.get("quality_valid_fraction", np.nan)
        out["COD_Status"] = frame_row.get("cod_status", "")

        keep = [
            "Specimen",
            "State",
            "Target_Strain_pct",
            "Frame",
            "Time_s",
            "Real_Strain_pct",
            "Stress_MPa",
            "Crack_ID",
            "Length_mm",
            "W_avg_um",
            "W_max_um",
            "Sample_Count",
            "Quality_Valid_Fraction",
            "COD_Status",
        ]
        for col in keep:
            if col not in out.columns:
                out[col] = np.nan
        return out[keep].sort_values(["State", "W_avg_um"], ascending=[True, False]).reset_index(drop=True)

    @staticmethod
    def _build_distribution_table(crack_tidy_df: pd.DataFrame) -> pd.DataFrame:
        if crack_tidy_df.empty:
            return pd.DataFrame(columns=["Specimen", "State", "Metric", "Value_um", "Crack_ID", "Frame", "Real_Strain_pct"])

        rows = []
        for _, row in crack_tidy_df.iterrows():
            for metric in ("W_avg_um", "W_max_um"):
                value = row.get(metric)
                if pd.notna(value):
                    rows.append(
                        {
                            "Specimen": row.get("Specimen"),
                            "State": row.get("State"),
                            "Metric": metric,
                            "Value_um": float(value),
                            "Crack_ID": row.get("Crack_ID"),
                            "Frame": row.get("Frame"),
                            "Real_Strain_pct": row.get("Real_Strain_pct"),
                        }
                    )
        return pd.DataFrame(rows)

    @staticmethod
    def _origin_curve_table(frame_df: pd.DataFrame) -> pd.DataFrame:
        cols = [
            "Specimen",
            "Frame",
            "Time_s",
            "Normalized_Time",
            "Strain_pct",
            "Normalized_Strain",
            "Stress_MPa",
            "crack_count",
            "crack_spacing_mm",
            "W_avg_um",
            "W_99_um",
            "W_max_um",
            "quality_valid_fraction",
            "cod_status",
            "sync_status",
        ]
        return frame_df[[c for c in cols if c in frame_df.columns]].copy()

    @staticmethod
    def _build_specimen_summary(
        specimen: str, frame_df: pd.DataFrame, sat_row: pd.Series, ult_row: pd.Series, crack_tidy_df: pd.DataFrame
    ) -> pd.DataFrame:
        has_stress = "Stress_MPa" in frame_df.columns and not frame_df["Stress_MPa"].isna().all()
        first_frame = frame_df.iloc[0] if not frame_df.empty else pd.Series(dtype=object)
        max_strain = float(pd.to_numeric(frame_df.get("Strain_pct", pd.Series(dtype=float)), errors="coerce").max(skipna=True) or 0.0)
        max_crack_count = int(pd.to_numeric(frame_df.get("crack_count", pd.Series(dtype=float)), errors="coerce").max(skipna=True) or 0)

        return pd.DataFrame(
            [
                {
                    "Specimen": specimen,
                    "Frame_Count": int(len(frame_df)),
                    "Time_Start_s": float(frame_df["Time_s"].min()) if "Time_s" in frame_df else np.nan,
                    "Time_End_s": float(frame_df["Time_s"].max()) if "Time_s" in frame_df else np.nan,
                    "Max_Strain_pct": max_strain,
                    "UTS_Stress_MPa": float(ult_row.get("Stress_MPa", np.nan)) if has_stress else np.nan,
                    "Ultimate_Frame": int(ult_row.get("Frame", -1)) if pd.notna(ult_row.get("Frame", np.nan)) else np.nan,
                    "Ultimate_Strain_pct": float(ult_row.get("Strain_pct", np.nan)) if pd.notna(ult_row.get("Strain_pct", np.nan)) else np.nan,
                    "Ultimate_W_99_um": float(ult_row.get("W_99_um", np.nan)) if pd.notna(ult_row.get("W_99_um", np.nan)) else np.nan,
                    "Ultimate_W_max_um": float(ult_row.get("W_max_um", np.nan)) if pd.notna(ult_row.get("W_max_um", np.nan)) else np.nan,
                    "Saturated_Frame": int(sat_row.get("Frame", -1)) if pd.notna(sat_row.get("Frame", np.nan)) else np.nan,
                    "Saturated_Strain_pct": float(sat_row.get("Strain_pct", np.nan)) if pd.notna(sat_row.get("Strain_pct", np.nan)) else np.nan,
                    "Saturated_Crack_Count": int(sat_row.get("crack_count", max_crack_count)) if pd.notna(sat_row.get("crack_count", np.nan)) else max_crack_count,
                    "Saturated_Spacing_mm": float(sat_row.get("crack_spacing_mm", np.nan)) if pd.notna(sat_row.get("crack_spacing_mm", np.nan)) else np.nan,
                    "Saturated_W_avg_um": float(sat_row.get("W_avg_um", np.nan)) if pd.notna(sat_row.get("W_avg_um", np.nan)) else np.nan,
                    "Crack_Detail_Rows": int(len(crack_tidy_df)),
                    "Pixel_Size_mm_per_px": first_frame.get("pixel_size_mm", np.nan),
                    "Subset_Spacing_px": first_frame.get("subset_spacing_px", np.nan),
                    "DIC_Point_Spacing_mm": first_frame.get("dic_point_spacing_mm", np.nan),
                    "Metadata_Source": first_frame.get("metadata_source", ""),
                    "DIC_Time_Source": first_frame.get("dic_time_source", ""),
                    "Strain_Source": ult_row.get("strain_source", ""),
                    "Sync_Status": ult_row.get("sync_status", ""),
                    "Worst_COD_Status_Count": int((frame_df.get("cod_status", pd.Series(dtype=str)).astype(str) != "ok").sum()) if "cod_status" in frame_df else 0,
                    "Min_Quality_Valid_Fraction": float(frame_df["quality_valid_fraction"].min()) if "quality_valid_fraction" in frame_df else np.nan,
                }
            ]
        )

    @staticmethod
    def _build_qa_frame_status(frame_df: pd.DataFrame) -> pd.DataFrame:
        cols = [
            "Specimen",
            "Frame",
            "Time_s",
            "Strain_pct",
            "quality_valid_fraction",
            "quality_filter",
            "quality_map_present",
            "v_map_present",
            "cod_status",
            "cod_vector_mode",
            "cod_sample_count",
            "sync_status",
            "dic_time_source",
            "strain_source",
            "metadata_source",
            "pixel_size_mm",
            "subset_spacing_px",
            "dic_point_spacing_mm",
            "strain_threshold_used",
        ]
        return frame_df[[c for c in cols if c in frame_df.columns]].copy()

    @staticmethod
    def _sheet_guide_frame() -> pd.DataFrame:
        return pd.DataFrame(
            [
                {"Sheet": "01_Frame_Curves", "Use": "直接画时序图：strain/stress/crack count/spacing/COD。"},
                {"Sheet": "02_Target_States", "Use": "目标应变点摘要。没达到也会保留 not_reached 行。"},
                {"Sheet": "03_Distribution_Tidy", "Use": "长表分布数据，适合 Origin boxplot/violin/KDE。"},
                {"Sheet": "04_Crack_Tidy", "Use": "一行一条裂缝，带 State/Frame/strain。透视表友好。"},
                {"Sheet": "Statistics_Report", "Use": "完整逐帧表、QA、验证、单试件摘要。"},
            ]
        )

    @staticmethod
    def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
        safe_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(df)
        if safe_df.empty and len(safe_df.columns) == 0:
            safe_df = pd.DataFrame({"Notice": ["No data"]})
        safe_df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(idx)].width = max(10, min(max_len + 2, 42))

    def _update_batch_summary(self, summary_df: pd.DataFrame, target_df: pd.DataFrame) -> None:
        batch_path = self.out_dir / "_Batch_Summary.xlsx"
        try:
            if batch_path.exists():
                old_summary = pd.read_excel(batch_path, sheet_name="Specimen_Summary")
                old_targets = pd.read_excel(batch_path, sheet_name="Target_Summary")
            else:
                old_summary = pd.DataFrame()
                old_targets = pd.DataFrame()

            specimen = str(summary_df.iloc[0]["Specimen"]) if not summary_df.empty else ""
            if not old_summary.empty and "Specimen" in old_summary.columns:
                old_summary = old_summary[old_summary["Specimen"].astype(str) != specimen]
            if not old_targets.empty and "Specimen" in old_targets.columns:
                old_targets = old_targets[old_targets["Specimen"].astype(str) != specimen]

            merged_summary = pd.concat([old_summary, summary_df], ignore_index=True)
            merged_targets = pd.concat([old_targets, target_df], ignore_index=True)

            with pd.ExcelWriter(batch_path, engine="openpyxl") as writer:
                self._write_sheet(writer, merged_summary, "Specimen_Summary")
                self._write_sheet(writer, merged_targets, "Target_Summary")
        except Exception as exc:
            logger.warning("Batch summary export failed: %s", exc, exc_info=True)
            self.log_emitted.emit(f"⚠️ Batch summary export failed: {exc}")

    @staticmethod
    def _details_for_frame(results: list[Dict[str, Any]], frame: Any) -> pd.DataFrame:
        if pd.isna(frame):
            return pd.DataFrame()
        for r in results:
            if int(r["Frame"]) == int(frame):
                return r.get("per_crack_details", pd.DataFrame())
        return pd.DataFrame()

    @staticmethod
    def _qa_frame(df: pd.DataFrame) -> pd.DataFrame:
        rows = []
        for key in [
            "pixel_size_mm",
            "subset_spacing_px",
            "dic_point_spacing_mm",
            "metadata_source",
            "dic_time_source",
            "v_map_present",
            "quality_map_present",
            "quality_filter",
            "quality_valid_fraction",
            "cod_vector_mode",
            "cod_status",
            "sync_status",
            "strain_source",
        ]:
            if key in df.columns:
                vals = df[key].dropna()
                rows.append(
                    {
                        "Item": key,
                        "First": vals.iloc[0] if not vals.empty else np.nan,
                        "Min": vals.min() if not vals.empty and pd.api.types.is_numeric_dtype(vals) else np.nan,
                        "Max": vals.max() if not vals.empty and pd.api.types.is_numeric_dtype(vals) else np.nan,
                        "Unique_Count": int(vals.astype(str).nunique()) if not vals.empty else 0,
                    }
                )
        return pd.DataFrame(rows)

    def _validation_frame(self, mat_path: Path, df: pd.DataFrame) -> pd.DataFrame:
        validation = self.config.get("validation", {})
        candidates = []
        if validation.get("annotation_path"):
            candidates.append(Path(validation["annotation_path"]))
        candidates.extend(
            [
                mat_path.with_name(f"{mat_path.stem}_annotations.csv"),
                mat_path.with_name(f"{mat_path.stem}_annotation.csv"),
                mat_path.with_name(f"{mat_path.stem}_annotations.xlsx"),
            ]
        )
        ann_path = next((p for p in candidates if p.exists()), None)
        if ann_path is None:
            return pd.DataFrame(
                {
                    "Status": ["No annotation file found"],
                    "Expected": ["<specimen>_annotations.csv/xlsx with Frame and optional crack_count/W_avg_um/W_max_um"],
                }
            )

        ann = pd.read_excel(ann_path) if ann_path.suffix.lower() in {".xls", ".xlsx"} else pd.read_csv(ann_path)
        if "Frame" not in ann.columns:
            return pd.DataFrame({"Status": [f"Annotation file has no Frame column: {ann_path.name}"]})
        merged = pd.merge(df, ann, on="Frame", suffixes=("_calc", "_manual"))
        if merged.empty:
            return pd.DataFrame({"Status": [f"No matching frames with annotation file: {ann_path.name}"]})
        rows = [{"Status": "OK", "Annotation_File": str(ann_path), "Matched_Frames": int(len(merged))}]
        for metric in ("crack_count", "W_avg_um", "W_max_um"):
            calc_col = f"{metric}_calc"
            manual_col = f"{metric}_manual"
            if calc_col in merged.columns and manual_col in merged.columns:
                err = pd.to_numeric(merged[calc_col], errors="coerce") - pd.to_numeric(
                    merged[manual_col], errors="coerce"
                )
                rows.append(
                    {
                        "Metric": metric,
                        "MAE": float(np.nanmean(np.abs(err))),
                        "Bias": float(np.nanmean(err)),
                        "MaxAbsError": float(np.nanmax(np.abs(err))),
                    }
                )
        return pd.DataFrame(rows)

    def stop(self) -> None:
        self._is_running = False
